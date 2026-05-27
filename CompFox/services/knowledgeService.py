import logging
import uuid
from typing import List, Optional

from pydantic import BaseModel
from langchain_text_splitters import RecursiveCharacterTextSplitter

from CompFox.models.pojo.knowledgeDocPo import KnowledgeDocPo
from CompFox.models.pojo.knowledgeBo import KnowledgeImportTextBo, KnowledgeImportUrlBo, KnowledgeSearchBo
from CompFox.models.vdb.knowledgeDocVdb import KnowledgeDocVdb

logger = logging.getLogger(__name__)

CHUNK_SIZE = 500
CHUNK_OVERLAP = 80

# 中文递归切片器：段落 → 换行 → 句号 → 问号/感叹号 → 逗号/分号 → 空格 → 字符
_CHINESE_SEPARATORS = [
    "\n\n",
    "\n",
    "。",
    "！",
    "？",
    "；",
    "，",
    "、",
    " ",
    "",
]


def _get_splitter(chunk_size: int = CHUNK_SIZE, overlap: int = CHUNK_OVERLAP) -> RecursiveCharacterTextSplitter:
    return RecursiveCharacterTextSplitter(
        separators=_CHINESE_SEPARATORS,
        chunk_size=chunk_size,
        chunk_overlap=overlap,
        keep_separator=True,
    )


def _parse_txt(content: str) -> str:
    """TXT 文本直接返回（调用方已完成编码检测）"""
    return content


def _parse_pdf(file_path: str) -> str:
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(file_path)
        pages = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                pages.append(text)
        return '\n'.join(pages)
    except ImportError:
        raise ImportError("请安装 PyPDF2: pip install PyPDF2")
    except Exception as e:
        logger.error(f"PDF 解析失败：{e}")
        raise


def _parse_docx(file_path: str) -> str:
    try:
        from docx import Document
        doc = Document(file_path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return '\n'.join(paragraphs)
    except ImportError:
        raise ImportError("请安装 python-docx: pip install python-docx")
    except Exception as e:
        logger.error(f"DOCX 解析失败：{e}")
        raise


def _parse_excel(file_path: str) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        rows = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows.append(f'[Sheet: {sheet_name}]')
            for row in ws.iter_rows(values_only=True):
                row_text = ' | '.join(str(cell) for cell in row if cell is not None)
                if row_text.strip():
                    rows.append(row_text)
        return '\n'.join(rows)
    except ImportError:
        raise ImportError("请安装 openpyxl: pip install openpyxl")
    except Exception as e:
        logger.error(f"Excel 解析失败：{e}")
        raise


def _read_file_with_encoding(file_path: str) -> str:
    """自适应编码读取文本文件"""
    for enc in ('utf-8', 'gbk', 'gb2312', 'latin-1'):
        try:
            with open(file_path, 'r', encoding=enc) as f:
                return f.read()
        except (UnicodeDecodeError, UnicodeError):
            continue
    raise ValueError(f"无法识别文件编码: {file_path}")


def _smart_chunk(text: str, chunk_size: int = CHUNK_SIZE, overlap: int = CHUNK_OVERLAP) -> List[str]:
    """
    递归切片：段落边界 → 换行 → 句号 → 感叹号/问号 → 分号 → 逗号 → 顿号 → 空格 → 字符

    Args:
        text: 待分块文本
        chunk_size: 每块最大字符数
        overlap: 块间重叠字符数

    Returns:
        List[str]: 分块后的文本列表
    """
    if not text or not text.strip():
        return []

    splitter = _get_splitter(chunk_size, overlap)
    chunks = splitter.split_text(text)
    return chunks if chunks else [text]


class KnowledgeService(BaseModel):
    """合规知识库服务"""

    @staticmethod
    def import_text(bo: KnowledgeImportTextBo) -> dict:
        """
        导入纯文本到知识库

        Args:
            bo: 文本导入参数

        Returns:
            {'doc_id': str, 'doc_name': str, 'total_chunks': int, 'total_chars': int}
        """
        doc_id = str(uuid.uuid4())
        content = bo.content
        total_chars = len(content)

        # 分块
        chunk_texts = _smart_chunk(content)
        if not chunk_texts:
            raise ValueError("文本内容为空，无法分块")

        # 构建 VDB 分块对象
        vdb_chunks = []
        for i, chunk_text in enumerate(chunk_texts):
            vdb_chunk = KnowledgeDocVdb(
                doc_id=doc_id,
                doc_name=bo.doc_name,
                chunk_index=str(i),
                content=chunk_text,
                source_type=bo.source_type or 'raw_text',
                source_url='',
                tags=bo.tags or '',
                effective_date=bo.effective_date or '',
                version=bo.version or '',
            )
            vdb_chunks.append(vdb_chunk)

        # 批量插入 Milvus（含 embedding）
        result = KnowledgeDocVdb.batch_insert_chunks(vdb_chunks)

        if result['failed'] > 0 and result['inserted'] == 0:
            raise RuntimeError(f"所有分块插入 Milvus 失败，共 {result['failed']} 个")

        # 写入 MySQL 元数据
        try:
            doc_po = KnowledgeDocPo(
                doc_id=doc_id,
                doc_name=bo.doc_name,
                source_type=bo.source_type or 'raw_text',
                source_url='',
                total_chunks=len(chunk_texts),
                total_chars=total_chars,
                tags=bo.tags,
                effective_date=bo.effective_date,
                version=bo.version,
                status=1,
            )
            doc_po.save()
        except Exception as e:
            logger.warning(f"写入 MySQL 元数据失败（Milvus 已写入）：{e}")

        logger.info(f"文本导入完成：doc_id={doc_id}, chunks={len(chunk_texts)}, "
                     f"inserted={result['inserted']}, failed={result['failed']}")

        return {
            'doc_id': doc_id,
            'doc_name': bo.doc_name,
            'total_chunks': len(chunk_texts),
            'total_chars': total_chars,
            'inserted': result['inserted'],
            'failed': result['failed'],
        }

    @staticmethod
    def import_file(file_path: str, file_name: str = None,
                    tags: str = None, effective_date: str = None,
                    version: str = None) -> dict:
        """
        导入文件（PDF/DOCX/TXT/Excel）到知识库

        Args:
            file_path: 文件路径
            file_name: 文档名称（可选，默认使用文件名）
            tags: 标签
            effective_date: 生效日期
            version: 版本号

        Returns:
            {'doc_id': str, 'doc_name': str, 'total_chunks': int, 'total_chars': int}
        """
        import os
        if not file_name:
            file_name = os.path.basename(file_path)

        ext = os.path.splitext(file_path)[1].lower()

        if ext == '.txt':
            content = _read_file_with_encoding(file_path)
            source_type = 'txt'
        elif ext == '.pdf':
            content = _parse_pdf(file_path)
            source_type = 'pdf'
        elif ext == '.docx':
            content = _parse_docx(file_path)
            source_type = 'docx'
        elif ext in ('.xlsx', '.xls'):
            content = _parse_excel(file_path)
            source_type = 'excel'
        else:
            raise ValueError(f"不支持的文件类型：{ext}")

        bo = KnowledgeImportTextBo(
            doc_name=file_name,
            content=content,
            source_type=source_type,
            tags=tags,
            effective_date=effective_date,
            version=version,
        )
        return KnowledgeService.import_text(bo)

    @staticmethod
    def import_url(bo: KnowledgeImportUrlBo) -> dict:
        """
        从 URL 抓取并导入知识库

        Args:
            bo: URL 导入参数

        Returns:
            {'doc_id': str, 'doc_name': str, 'total_chunks': int, 'total_chars': int}
        """
        import requests

        try:
            resp = requests.get(bo.url, timeout=30, headers={
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            })
            resp.raise_for_status()
        except Exception as e:
            logger.error(f"URL 抓取失败：{e}")
            raise RuntimeError(f"URL 抓取失败：{e}")

        # 推断来源类型
        url_lower = bo.url.lower()
        if bo.source_type:
            source_type = bo.source_type
        elif url_lower.endswith('.pdf'):
            source_type = 'pdf'
        elif url_lower.endswith('.docx'):
            source_type = 'docx'
        elif url_lower.endswith('.txt'):
            source_type = 'txt'
        elif url_lower.endswith('.xlsx') or url_lower.endswith('.xls'):
            source_type = 'excel'
        else:
            source_type = 'url'

        doc_name = bo.doc_name or bo.url.rsplit('/', 1)[-1] or 'web_page'

        # 简单提取文本（去除 HTML 标签）
        content = resp.text
        if source_type == 'url':
            import re as _re
            content = _re.sub(r'<script[^>]*>.*?</script>', '', content, flags=_re.DOTALL | _re.IGNORECASE)
            content = _re.sub(r'<style[^>]*>.*?</style>', '', content, flags=_re.DOTALL | _re.IGNORECASE)
            content = _re.sub(r'<[^>]+>', '', content)
            content = _re.sub(r'\n\s*\n', '\n\n', content)
            content = _re.sub(r'[ \t]+', ' ', content)

        if not content or not content.strip():
            raise ValueError("URL 内容为空")

        text_bo = KnowledgeImportTextBo(
            doc_name=doc_name,
            content=content,
            source_type=source_type,
            tags=bo.tags,
            effective_date=bo.effective_date,
            version=bo.version,
        )
        result = KnowledgeService.import_text(text_bo)
        result['source_url'] = bo.url
        return result

    @staticmethod
    def retrieve_context(query: str, limit: int = 5,
                         source_type: str = None,
                         tags: str = None) -> str:
        """
        检索知识库，返回格式化的法规上下文文本供 Prompt 注入。
        Milvus 不可用时返回空字符串，系统降级为无 RAG 模式。

        Args:
            query: 检索查询文本
            limit: 返回结果数量上限
            source_type: 按来源类型过滤
            tags: 按标签过滤

        Returns:
            str: 格式化的法规上下文文本
        """
        return KnowledgeDocVdb.retrieve_context(query, limit, source_type, tags)

    @staticmethod
    def search(bo: KnowledgeSearchBo) -> List[dict]:
        """检索知识库"""
        return KnowledgeDocVdb.search_knowledge(
            query=bo.query,
            limit=bo.limit,
            source_type=bo.source_type,
            tags=bo.tags,
        )

    @staticmethod
    def list_documents() -> List[dict]:
        """列出知识库文档清单（从 MySQL）"""
        docs = KnowledgeDocPo.list_active()
        result = []
        for doc in docs:
            result.append({
                'doc_id': doc.doc_id,
                'doc_name': doc.doc_name,
                'source_type': doc.source_type,
                'source_url': doc.source_url,
                'total_chunks': doc.total_chunks,
                'total_chars': doc.total_chars,
                'tags': doc.tags,
                'effective_date': doc.effective_date,
                'version': doc.version,
                'created_at': doc.created_at.isoformat() if doc.created_at else None,
                'updated_at': doc.updated_at.isoformat() if doc.updated_at else None,
            })
        return result

    @staticmethod
    def delete_document(doc_id: str) -> bool:
        """删除知识库文档（Milvus + MySQL）"""
        vdb_ok = KnowledgeDocVdb.delete_by_doc_id(doc_id)
        mysql_ok = KnowledgeDocPo.soft_delete(doc_id)
        if not vdb_ok and not mysql_ok:
            return False
        return True


knowledge_service = KnowledgeService()


def get_knowledge_service():
    return knowledge_service
